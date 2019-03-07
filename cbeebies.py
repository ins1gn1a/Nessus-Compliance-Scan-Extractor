#!/usr/bin/env python3

import sys
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from datetime import datetime

currenttime = datetime.today().strftime('%Y-%m-%d_%H-%M')

tree = ET.parse(sys.argv[1])
root = tree.getroot()
    
# Excel Workbook
dest_filename = 'ComplianceScan-' + currenttime + '.xlsx'
wb = Workbook()

# Worksheet : Passed
ws1 = wb.active
ws1.title = "Passed"
ws2 = wb.create_sheet(title="Failed")
ws3 = wb.create_sheet(title="Warning")

header_list = [['Result','Host','Compliance Check','Overview','Details','Solution','References']]
ws1.append(header_list[0])
ws2.append(header_list[0])
ws3.append(header_list[0])

for ReportHost in root.findall("./Report/ReportHost"): # Loops through ReportHosts
    passed_host=passed_name=passed_description=passed_actual=passed_info=passed_result=passed_solution=passed_reference = ""
    z = []
    passed_host = (ReportHost.attrib['name'])
    
    previousResults = ([])

    for i in ReportHost: # Loops through reportItems
        for x in i: # Loops through Tags in reportItems
            if x.tag == "{http://www.nessus.org/cm}compliance-result": 
                passed_result = x.text
                pass
            if x.tag == "{http://www.nessus.org/cm}compliance-check-name": 
                passed_name = x.text
                pass
            if x.tag == "{http://www.nessus.org/cm}compliance-info": 
                passed_info = x.text
                pass
            if x.tag == "{http://www.nessus.org/cm}compliance-actual-value": 
                passed_actual = x.text
                pass
            if x.tag == "{http://www.nessus.org/cm}compliance-solution": 
                passed_solution = x.text
                pass
            if x.tag == "{http://www.nessus.org/cm}compliance-reference": 
                passed_reference = x.text
                pass

        if passed_result == 'PASSED':
            if previousResults != ([passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference]):
                ws1.append([passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference])
            previousResults = [passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference]
        
        if passed_result == 'FAILED':
            if previousResults != ([passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference]):
                ws2.append([passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference])
            previousResults = [passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference]  

        if passed_result == 'WARNING':
            if previousResults != ([passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference]):
                ws3.append([passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference])
            previousResults = [passed_result,passed_host,passed_name,passed_info,passed_actual,passed_solution,passed_reference]              
            
wb.save(filename = dest_filename)
